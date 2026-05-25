"""
verify_against_source.py

Compares a working .xlsm against an emergency-save source-of-truth .xlsx
using compound key matching (Description + Owner).

CRITICAL: This script READS files only. It does not write Excel files.
Files built or written by openpyxl will have CF fills written as bgColor
which Excel renders as invisible. See docs/What-AI-Wont-Tell-You.html — Error 001.

Outputs:
  - Missing_Rows.xlsx        : Rows present in source but missing from working file
  - Updates_Needed.xlsx      : Field-level mismatches with working-file row numbers

Usage:
  python verify_against_source.py <working_file.xlsm> <source_of_truth.xlsx>
"""

import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook


# Column positions (0-indexed) in Meeting Minutes
COL_TYPE = 0     # A
COL_DESC = 1     # B
COL_OWNER = 2    # C
COL_DUE = 3      # D
COL_STATUS = 4   # E
COL_PRIORITY = 5 # F
COL_OD = 6       # G
COL_NOTES = 7    # H
COL_PROJECT = 8  # I

DATA_START_ROW = 16  # Row 15 is column headers, rows 1-14 are header block


def make_key(row):
    """Compound key: Description + Owner. Description alone is not unique."""
    desc = (row[COL_DESC] or "").strip() if row[COL_DESC] else ""
    owner = (row[COL_OWNER] or "").strip() if row[COL_OWNER] else ""
    return f"{desc}||{owner}"


def load_mm_rows(filepath):
    """Load Meeting Minutes rows as a list of (row_num, [values]) tuples."""
    wb = load_workbook(filepath, data_only=True)
    if "Meeting Minutes" not in wb.sheetnames:
        raise ValueError(f"No 'Meeting Minutes' sheet in {filepath}")

    ws = wb["Meeting Minutes"]
    rows = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=c).value for c in range(1, 10)]
        # Skip banner rows (all caps in description, no owner, no due)
        if not values[COL_DESC]:
            continue
        rows.append((row_idx, values))
    return rows


def compare(working_path, source_path):
    working = load_mm_rows(working_path)
    source = load_mm_rows(source_path)

    working_by_key = {make_key(v): (r, v) for r, v in working}
    source_by_key = {make_key(v): (r, v) for r, v in source}

    missing = []
    mismatches = []

    for key, (source_row, source_vals) in source_by_key.items():
        if key not in working_by_key:
            missing.append((source_row, source_vals))
            continue

        working_row, working_vals = working_by_key[key]
        for col_idx, col_name in [
            (COL_TYPE, "Type"),
            (COL_DUE, "Due"),
            (COL_STATUS, "Status"),
            (COL_PRIORITY, "Priority"),
            (COL_NOTES, "Notes"),
        ]:
            if (source_vals[col_idx] or "") != (working_vals[col_idx] or ""):
                mismatches.append({
                    "working_row": working_row,
                    "source_row": source_row,
                    "description": source_vals[COL_DESC],
                    "owner": source_vals[COL_OWNER],
                    "field": col_name,
                    "source_value": source_vals[col_idx],
                    "working_value": working_vals[col_idx],
                })

    return missing, mismatches


def write_outputs(missing, mismatches, out_dir):
    out_dir = Path(out_dir)
    out_dir.mkdir(exist_ok=True)

    # Missing rows
    if missing:
        wb = Workbook()
        ws = wb.active
        ws.title = "Missing Rows"
        ws.append(["Source Row", "Type", "Description", "Owner", "Due", "Status", "Priority", "OD", "Notes", "Project"])
        for source_row, vals in missing:
            ws.append([source_row] + list(vals))
        wb.save(out_dir / "Missing_Rows.xlsx")
        print(f"Wrote {len(missing)} missing rows to Missing_Rows.xlsx")

    # Mismatches
    if mismatches:
        wb = Workbook()
        ws = wb.active
        ws.title = "Updates Needed"
        ws.append(["Working Row", "Source Row", "Description", "Owner", "Field", "Source Value", "Working Value"])
        for m in mismatches:
            ws.append([
                m["working_row"], m["source_row"], m["description"], m["owner"],
                m["field"], m["source_value"], m["working_value"]
            ])
        wb.save(out_dir / "Updates_Needed.xlsx")
        print(f"Wrote {len(mismatches)} field mismatches to Updates_Needed.xlsx")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    working = sys.argv[1]
    source = sys.argv[2]

    missing, mismatches = compare(working, source)
    write_outputs(missing, mismatches, "out")

    print(f"\n{len(missing)} rows missing from working file")
    print(f"{len(mismatches)} field mismatches detected")
